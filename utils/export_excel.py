"""
Export extracted paper evidence to a styled Excel file
matching the REST Table 2 format.
"""

from typing import List
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from core.schemas import PaperEvidence


# Column definitions matching REST Table 2
COLUMNS = [
    ("Article Reference", 45),
    ("Country", 15),
    ("Study Design", 25),
    ("Population", 30),
    ("Setting", 40),
    ("Peer Reviewed", 14),
    ("Intervention", 14),
    ("Primary Results", 55),
    ("Additional Findings", 45),
]


def export_to_excel(
    papers: List[PaperEvidence],
    output_path: str = "evidence_table.xlsx",
) -> Path:
    """
    Write a list of PaperEvidence objects to a styled Excel workbook.

    Returns the Path to the saved file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Table 2 – Evidence Summary"

    # ── Styles ─────────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="217346", end_color="217346", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_font = Font(name="Calibri", size=10)
    cell_alignment = Alignment(vertical="top", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    alt_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # ── Headers ────────────────────────────────────────────────────────
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = col_width

    ws.row_dimensions[1].height = 30

    # ── Data rows ──────────────────────────────────────────────────────
    for row_idx, paper in enumerate(papers, start=2):
        values = [
            paper.article_reference,
            paper.country,
            paper.study_type,
            paper.population,
            paper.setting,
            paper.peer_reviewed,
            paper.intervention,
            paper.primary_results,
            paper.additional_findings or "",
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Alternate row shading
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # ── Freeze header row ──────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ────────────────────────────────────────────────────
    if papers:
        last_col_letter = ws.cell(row=1, column=len(COLUMNS)).column_letter
        ws.auto_filter.ref = f"A1:{last_col_letter}{len(papers) + 1}"

    # ── Save ───────────────────────────────────────────────────────────
    out = Path(output_path)
    wb.save(out)
    print(f"\n✅ Evidence table saved to: {out.resolve()}")
    return out
